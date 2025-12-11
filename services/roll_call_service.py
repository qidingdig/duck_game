#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Roll call service: handles SQLite storage, student roster, leave records, and roll-call logs.

使用Repository模式实现，提供清晰的架构和易于扩展的设计。
"""

from __future__ import annotations

import csv
import json
import os
import time
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None

from data.database_interface import DatabaseInterface
from data.sqlite_database import SQLiteDatabase
from data.repositories import (
    StudentRepository,
    StudentLeaveRepository,
    RollCallRepository,
    RollCallRecordRepository,
    SQLiteStudentRepository,
    SQLiteStudentLeaveRepository,
    SQLiteRollCallRepository,
    SQLiteRollCallRecordRepository,
)
from data.models import Student, StudentLeave, RollCall, RollCallRecord
from data.database_migration import DatabaseMigration


DB_DEFAULT_PATH = os.path.join("data", "roll_call.db")

# 状态映射常量
STATUS_MAP = {
    "present": "出勤",
    "absent": "旷课",
    "leave": "请假",
    "late": "迟到",
}

# 学生信息列名映射配置
STUDENT_COLUMN_MAPPING = {
    'student_id': [
        '学号', 'student_id', 'studentid', 'id'
    ],
    'name': [
        '姓名', 'name'
    ],
    'nickname': [
        '昵称', 'nickname', 'nick'
    ],
    'photo_path': [
        '照片', 'photo', 'photo_path', '头像'
    ],
}


class RollCallService:
    """
    点名服务 - 使用Repository模式实现
    
    设计特点：
    1. Repository模式分离数据访问逻辑
    2. 数据库抽象接口，易于切换数据库
    3. 模型层清晰的实体定义
    4. 支持数据库迁移
    """
    
    def __init__(self, db_path: str = DB_DEFAULT_PATH, db: Optional[DatabaseInterface] = None):
        """
        初始化服务
        
        Args:
            db_path: 数据库文件路径
            db: 可选的数据库接口（用于测试或自定义实现）
        """
        # 初始化数据库
        if db is None:
            os.makedirs(os.path.dirname(db_path), exist_ok=True)
            self.db: DatabaseInterface = SQLiteDatabase(db_path)
        else:
            self.db = db
        
        # 初始化Repository
        self.student_repo: StudentRepository = SQLiteStudentRepository(self.db)
        self.leave_repo: StudentLeaveRepository = SQLiteStudentLeaveRepository(self.db)
        self.roll_call_repo: RollCallRepository = SQLiteRollCallRepository(self.db)
        self.record_repo: RollCallRecordRepository = SQLiteRollCallRecordRepository(self.db)
        
        # 执行数据库迁移
        migration = DatabaseMigration(self.db, db_path)
        migration.migrate()
        migration.seed_sample_data()
    
    # --------------------------------------------------------------------- #
    # Student operations
    # --------------------------------------------------------------------- #
    def list_students(self) -> List[Student]:
        """列出所有学生"""
        return self.student_repo.find_all()
    
    def select_students(
        self,
        strategy: str,
        count: Optional[int] = None,
    ) -> List[Student]:
        """
        根据策略选择学生
        
        Args:
            strategy: 选择策略 ("random", "most_absent", "least_called")
            count: 选择数量
        
        Returns:
            学生列表
        """
        students = self.student_repo.find_all()
        
        if strategy == "most_absent":
            students.sort(key=lambda s: s.cut_count, reverse=True)
        elif strategy == "least_called":
            students.sort(key=lambda s: s.called_count)
        else:
            import random
            random.shuffle(students)
        
        if count and count > 0:
            # 如果请求的数量大于等于总人数，返回所有学生
            # 否则返回指定数量的学生
            students = students[:min(count, len(students))]
        
        return students
    
    def has_leave(self, student_id: str, session_code: str) -> bool:
        """检查学生是否有请假记录"""
        return self.leave_repo.has_leave(student_id, session_code)
    
    # --------------------------------------------------------------------- #
    # Roll call operations
    # --------------------------------------------------------------------- #
    def start_roll_call(
        self,
        session_code: str,
        mode: str,
        strategy: str,
        selected_count: int,
    ) -> int:
        """
        开始点名会话
        
        Returns:
            点名会话ID
        """
        roll_call = RollCall(
            session_code=session_code,
            mode=mode,
            strategy=strategy,
            selected_count=selected_count,
            started_at=time.strftime("%Y-%m-%d %H:%M:%S"),
        )
        return self.roll_call_repo.create(roll_call)
    
    def insert_record(
        self,
        roll_call_id: int,
        student_id: str,
        order_index: int,
        status: str,
        note: str = "",
    ) -> int:
        """
        插入点名记录
        
        Returns:
            记录ID
        """
        # 获取学生信息，保存姓名快照（确保历史记录准确）
        student = self.student_repo.find_by_id(student_id)
        student_name = student.name if student else None
        
        record = RollCallRecord(
            roll_call_id=roll_call_id,
            student_id=student_id,
            student_name=student_name,  # 保存点名时的学生姓名快照
            order_index=order_index,
            status=status,
            called_time=time.strftime("%Y-%m-%d %H:%M:%S"),
            note=note,
        )
        record_id = self.record_repo.create(record)
        
        # 更新学生统计
        self._update_statistics(student_id, status, delta=1)
        
        return record_id
    
    def update_record_status(
        self,
        record_id: int,
        new_status: str,
        enforce_within_minutes: Optional[int] = None,
    ) -> bool:
        """
        更新记录状态
        
        Args:
            record_id: 记录ID
            new_status: 新状态
            enforce_within_minutes: 限制在多少分钟内可以修改
        
        Returns:
            是否成功
        """
        record = self.record_repo.find_by_id(record_id)
        if not record:
            return False
        
        old_status = record.status
        
        # 如果要改为迟到，必须满足：原状态是旷课
        if new_status == "late":
            if old_status != "absent":
                return False
        
        # 检查时间限制
        if enforce_within_minutes is not None:
            called_ts = time.mktime(time.strptime(record.called_time, "%Y-%m-%d %H:%M:%S"))
            if time.time() - called_ts > enforce_within_minutes * 60:
                return False
        
        updated_time = time.strftime("%Y-%m-%d %H:%M:%S")
        
        success = self.record_repo.update_status(record_id, new_status, updated_time)
        
        if success:
            # 调整统计信息
            self._adjust_statistics_for_change(record_id, old_status, new_status)
        
        return success
    
    def get_roll_call_summary(self, session_code: Optional[str] = None) -> List[Dict[str, Any]]:
        """获取点名摘要"""
        # 使用保存的student_name快照，确保历史记录准确
        query = """
            SELECT r.session_code, r.started_at, rr.status, rr.student_id, 
                   COALESCE(rr.student_name, s.name) as name
            FROM roll_call_records rr
            JOIN roll_calls r ON rr.roll_call_id = r.id
            LEFT JOIN students s ON rr.student_id = s.student_id
        """
        params: tuple = ()
        if session_code:
            query += " WHERE r.session_code = ?"
            params = (session_code,)
        query += " ORDER BY r.started_at DESC, rr.order_index ASC"
        
        rows = self.db.fetch_all(query, params)
        
        summary = []
        for row in rows:
            summary.append({
                "session_code": row[0],
                "started_at": row[1],
                "status": row[2],
                "student_id": row[3],
                "name": row[4],  # 优先使用保存的student_name，如果没有则使用当前学生表的name
            })
        
        return summary
    
    def list_all_sessions(self) -> List[Dict[str, Any]]:
        """
        列出所有点名会话
        
        Returns:
            会话列表，包含session_code、started_at、mode、strategy、selected_count等信息
        """
        query = """
            SELECT DISTINCT r.session_code, r.started_at, r.mode, r.strategy, r.selected_count,
                   COUNT(rr.id) as record_count
            FROM roll_calls r
            LEFT JOIN roll_call_records rr ON r.id = rr.roll_call_id
            GROUP BY r.id, r.session_code, r.started_at, r.mode, r.strategy, r.selected_count
            ORDER BY r.started_at DESC
        """
        rows = self.db.fetch_all(query)
        
        sessions = []
        for row in rows:
            sessions.append({
                "session_code": row[0],
                "started_at": row[1],
                "mode": row[2],
                "strategy": row[3],
                "selected_count": row[4],
                "record_count": row[5] or 0,
            })
        
        return sessions
    
    def get_session_details(self, session_code: str) -> List[Dict[str, Any]]:
        """
        获取指定会话的详细记录
        
        Args:
            session_code: 会话代码
            
        Returns:
            详细记录列表
        """
        query = """
            SELECT 
                rr.id,
                rr.student_id,
                COALESCE(rr.student_name, s.name) as name,
                rr.order_index,
                rr.status,
                rr.called_time,
                rr.updated_time,
                rr.note
            FROM roll_call_records rr
            JOIN roll_calls r ON rr.roll_call_id = r.id
            LEFT JOIN students s ON rr.student_id = s.student_id
            WHERE r.session_code = ?
            ORDER BY rr.order_index ASC
        """
        rows = self.db.fetch_all(query, (session_code,))
        
        details = []
        for row in rows:
            details.append({
                "id": row[0],
                "student_id": row[1],
                "name": row[2],  # 优先使用保存的student_name，如果没有则使用当前学生表的name
                "order_index": row[3],
                "status": row[4],
                "called_time": row[5],
                "updated_time": row[6],
                "note": row[7] or "",
            })
        
        return details
    
    def delete_session(self, session_code: str) -> bool:
        """
        删除整个点名会话及其所有记录
        
        Args:
            session_code: 会话代码
            
        Returns:
            是否删除成功
        """
        # 获取会话信息
        roll_call = self.roll_call_repo.find_by_session_code(session_code)
        if not roll_call or not roll_call.id:
            return False
        
        # 获取要删除的所有记录，用于更新统计
        records = self.record_repo.find_by_roll_call_id(roll_call.id)
        
        # 删除会话（会自动删除所有相关记录）
        success = self.roll_call_repo.delete(roll_call.id)
        
        if success:
            # 更新学生统计信息
            for record in records:
                self._update_statistics(record.student_id, record.status, delta=-1)
        
        return success
    
    def delete_sessions(self, session_codes: List[str]) -> int:
        """
        批量删除点名会话
        
        Args:
            session_codes: 会话代码列表
            
        Returns:
            删除的会话数
        """
        deleted_count = 0
        for session_code in session_codes:
            if self.delete_session(session_code):
                deleted_count += 1
        return deleted_count
    
    def get_latest_record(self, roll_call_id: int, student_id: str) -> Optional[Dict[str, Any]]:
        """获取最新记录"""
        record = self.record_repo.find_latest_by_roll_call_and_student(roll_call_id, student_id)
        if not record:
            return None
        
        return {
            "id": record.id,
            "status": record.status,
            "called_time": record.called_time,
            "updated_time": record.updated_time,
            "note": record.note,
        }
    
    # --------------------------------------------------------------------- #
    # Internal helpers
    # --------------------------------------------------------------------- #
    def _update_statistics(self, student_id: str, status: str, delta: int) -> None:
        """更新学生统计信息"""
        if status not in {"present", "leave", "absent", "late"}:
            return
        
        if status == "absent":
            self.student_repo.update_statistics(student_id, cut_delta=delta, called_delta=delta)
        else:
            self.student_repo.update_statistics(student_id, cut_delta=0, called_delta=delta)
    
    def _adjust_statistics_for_change(self, record_id: int, old_status: str, new_status: str) -> None:
        """调整统计信息（状态变更时）"""
        if old_status == new_status:
            return
        
        record = self.record_repo.find_by_id(record_id)
        if not record:
            return
        
        student_id = record.student_id
        
        # 撤销旧状态的影响
        self._update_statistics(student_id, old_status, delta=-1)
        
        # 应用新状态的影响
        self._update_statistics(student_id, new_status, delta=1)
    
    def export_to_csv(self, session_code: Optional[str] = None, output_path: str = "roll_call_export.csv") -> str:
        """
        导出点名记录为CSV格式
        
        Args:
            session_code: 可选，指定会话代码，None表示导出所有
            output_path: 输出文件路径
            
        Returns:
            输出文件路径
        """
        import csv
        
        # 如果指定了会话代码，使用旧格式（一行一条记录）
        if session_code:
            records = self.get_roll_call_summary(session_code)
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(['会话代码', '开始时间', '学号', '姓名', '状态'])
                for record in records:
                    status = STATUS_MAP.get(record['status'], record['status'])
                    writer.writerow([
                        record['session_code'],
                        record['started_at'],
                        record['student_id'],
                        record['name'],
                        status,
                    ])
            return output_path
        
        # 导出所有会话时，使用新格式（一行为一个学生，每列是一次点名）
        # 获取所有学生
        all_students = self.student_repo.find_all()
        students_dict = {s.student_id: s for s in all_students}
        
        # 获取所有会话（按时间排序）
        all_sessions = self.list_all_sessions()
        # 按开始时间升序排列
        all_sessions.sort(key=lambda x: x['started_at'])
        
        # 构建学生-会话状态矩阵
        # 查询所有记录
        query = """
            SELECT r.session_code, rr.student_id, rr.status
            FROM roll_call_records rr
            JOIN roll_calls r ON rr.roll_call_id = r.id
            ORDER BY r.started_at ASC, rr.student_id ASC
        """
        rows = self.db.fetch_all(query)
        
        # 构建状态字典：{student_id: {session_code: status}}
        student_status_map = {}
        for row in rows:
            session_code, student_id, status = row
            if student_id not in student_status_map:
                student_status_map[student_id] = {}
            student_status_map[student_id][session_code] = status
        
        # 写入CSV
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # 表头：学号、姓名、会话1、会话2、...
            header = ['学号', '姓名']
            for session in all_sessions:
                # 使用会话代码和开始时间作为列名
                header.append(f"{session['session_code']}\n{session['started_at']}")
            writer.writerow(header)
            
            # 写入每个学生的数据
            for student in all_students:
                row = [student.student_id, student.name]
                for session in all_sessions:
                    session_code = session['session_code']
                    if student.student_id in student_status_map and session_code in student_status_map[student.student_id]:
                        status = student_status_map[student.student_id][session_code]
                        row.append(STATUS_MAP.get(status, status))
                    else:
                        row.append("")  # 没有点到该学生，留空
                writer.writerow(row)
        
        return output_path
    
    def export_to_excel(self, session_code: Optional[str] = None, output_path: str = "roll_call_export.xlsx") -> str:
        """
        导出点名记录为XLSX格式（Excel）
        
        Args:
            session_code: 可选，指定会话代码，None表示导出所有
            output_path: 输出文件路径（默认.xlsx扩展名）
            
        Returns:
            输出文件路径
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError("需要安装openpyxl库: pip install openpyxl")
        
        # 如果指定了会话代码，使用旧格式（一行一条记录）
        if session_code:
            records = self.get_roll_call_summary(session_code)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "点名记录"
            
            headers = ['会话代码', '开始时间', '学号', '姓名', '状态']
            ws.append(headers)
            
            for record in records:
                status = STATUS_MAP.get(record['status'], record['status'])
                ws.append([
                    record['session_code'],
                    record['started_at'],
                    record['student_id'],
                    record['name'],
                    status,
                ])
            
            wb.save(output_path)
            return output_path
        
        # 导出所有会话时，使用新格式（一行为一个学生，每列是一次点名）
        # 获取所有学生
        all_students = self.student_repo.find_all()
        
        # 获取所有会话（按时间排序）
        all_sessions = self.list_all_sessions()
        all_sessions.sort(key=lambda x: x['started_at'])
        
        # 构建学生-会话状态矩阵
        query = """
            SELECT r.session_code, rr.student_id, rr.status
            FROM roll_call_records rr
            JOIN roll_calls r ON rr.roll_call_id = r.id
            ORDER BY r.started_at ASC, rr.student_id ASC
        """
        rows = self.db.fetch_all(query)
        
        # 构建状态字典：{student_id: {session_code: status}}
        student_status_map = {}
        for row in rows:
            session_code, student_id, status = row
            if student_id not in student_status_map:
                student_status_map[student_id] = {}
            student_status_map[student_id][session_code] = status
        
        # 创建Excel文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "点名记录"
        
        # 表头：学号、姓名、会话1、会话2、...
        header = ['学号', '姓名']
        for session in all_sessions:
            # 使用会话代码和开始时间作为列名
            header.append(f"{session['session_code']}\n{session['started_at']}")
        ws.append(header)
        
        # 写入每个学生的数据
        for student in all_students:
            row = [student.student_id, student.name]
            for session in all_sessions:
                session_code = session['session_code']
                if student.student_id in student_status_map and session_code in student_status_map[student.student_id]:
                    status = student_status_map[student.student_id][session_code]
                    row.append(STATUS_MAP.get(status, status))
                else:
                    row.append("")  # 没有点到该学生，留空
            ws.append(row)
        
        wb.save(output_path)
        return output_path
    
    def import_students_from_file(self, file_path: str, update_existing: bool = True) -> Dict[str, Any]:
        """
        从文件导入学生信息
        
        支持格式：
        - CSV: 必须包含学号(student_id)和姓名(name)列，可选nickname, photo_path列
        - Excel: 同上
        - JSON: 数组格式，每个元素包含student_id, name等字段
        
        更新策略：
        - 通过 student_id 判断学生是否已存在
        - 如果 update_existing=True（默认）：
          * 已存在：更新基本信息（name, nickname, photo_path），保留统计信息（cut_count, called_count）
          * 不存在：新增学生
        - 如果 update_existing=False：
          * 已存在：跳过该记录，计入 skipped
          * 不存在：新增学生
        
        Args:
            file_path: 文件路径
            update_existing: 是否更新已存在的学生（默认True）
            
        Returns:
            包含导入结果的字典：
            {
                'success': bool,
                'total': int,      # 总记录数
                'imported': int,   # 新增数量
                'updated': int,    # 更新数量（仅在update_existing=True时）
                'skipped': int,    # 跳过数量（无效记录或已存在且update_existing=False）
                'errors': List[str],  # 错误信息列表（真正的错误，如数据格式错误、必需字段缺失等）
                'warnings': List[str]  # 警告信息列表（如已存在且跳过更新）
            }
        """
        result = {
            'success': False,
            'total': 0,
            'imported': 0,
            'updated': 0,
            'skipped': 0,
            'errors': [],
            'warnings': []
        }
        
        if not os.path.exists(file_path):
            result['errors'].append(f"文件不存在: {file_path}")
            return result
        
        from utils.config import Config
        config = Config()
        
        file_ext = os.path.splitext(file_path)[1].lower()
        
        try:
            # 根据文件类型解析数据
            if file_ext not in config.STUDENT_IMPORT_SUPPORTED_FORMATS:
                result['errors'].append(f"不支持的文件格式: {file_ext}")
                return result
            
            if file_ext == '.csv':
                students_data = self._parse_csv(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                if not HAS_OPENPYXL:
                    result['errors'].append("需要安装openpyxl库: pip install openpyxl")
                    return result
                students_data = self._parse_excel(file_path)
            elif file_ext == '.json':
                students_data = self._parse_json(file_path)
            else:
                result['errors'].append(f"不支持的文件格式: {file_ext}")
                return result
            
            result['total'] = len(students_data)
            
            # 导入数据
            for idx, student_data in enumerate(students_data, 1):
                process_result = self._process_student_record(student_data, idx, update_existing)
                result['imported'] += process_result['imported']
                result['updated'] += process_result['updated']
                result['skipped'] += process_result['skipped']
                result['errors'].extend(process_result['errors'])
                result['warnings'].extend(process_result['warnings'])
            
            result['success'] = True
            
        except Exception as e:
            result['errors'].append(f"导入过程出错: {str(e)}")
            import traceback
            result['errors'].append(traceback.format_exc())
        
        return result
    
    def _validate_student_data(self, student_data: Dict[str, Any], row_index: int) -> Tuple[Optional[str], Optional[str], List[str]]:
        """
        验证学生数据
        
        Returns:
            (student_id, name, errors) 元组
        """
        errors = []
        
        # 验证必需字段
        if not student_data.get('student_id') or not student_data.get('name'):
            errors.append(f"第{row_index}行: 缺少必需字段(student_id或name)")
            return None, None, errors
        
        student_id = str(student_data['student_id']).strip()
        name = str(student_data['name']).strip()
        
        if not student_id or not name:
            errors.append(f"第{row_index}行: 学号或姓名为空")
            return None, None, errors
        
        return student_id, name, errors
    
    def _process_student_record(
        self, 
        student_data: Dict[str, Any], 
        row_index: int, 
        update_existing: bool
    ) -> Dict[str, Any]:
        """
        处理单条学生记录
        
        Returns:
            包含 imported, updated, skipped, errors, warnings 的字典
        """
        result = {
            'imported': 0,
            'updated': 0,
            'skipped': 0,
            'errors': [],
            'warnings': []
        }
        
        try:
            # 验证数据
            student_id, name, validation_errors = self._validate_student_data(student_data, row_index)
            if validation_errors:
                result['skipped'] = 1
                result['errors'].extend(validation_errors)
                return result
            
            # 检查学生是否已存在
            existing_student = self.student_repo.find_by_id(student_id)
            
            # 如果学生已存在且不允许更新，则跳过
            if existing_student and not update_existing:
                result['skipped'] = 1
                result['warnings'].append(f"第{row_index}行: 学生 {student_id} ({name}) 已存在，跳过更新")
                return result
            
            # 创建或更新学生对象
            student = self._create_or_update_student(student_data, student_id, name, existing_student, row_index, result)
            
            # 保存学生
            self.student_repo.save(student)
            
            if existing_student:
                result['updated'] = 1
            else:
                result['imported'] = 1
                
        except Exception as e:
            result['skipped'] = 1
            result['errors'].append(f"第{row_index}行处理失败: {str(e)}")
        
        return result
    
    def _create_or_update_student(
        self,
        student_data: Dict[str, Any],
        student_id: str,
        name: str,
        existing_student: Optional[Student],
        row_index: int,
        result: Dict[str, Any]
    ) -> Student:
        """创建或更新学生对象"""
        if existing_student:
            # 检查姓名是否改变，给出警告提示
            if existing_student.name != name:
                result['warnings'].append(
                    f"第{row_index}行: 学生 {student_id} 的姓名从 '{existing_student.name}' 更新为 '{name}'。"
                    f"历史点名记录中的姓名已保存快照，不会改变。"
                )
            # 更新学生信息，保留统计信息
            return Student(
                student_id=student_id,
                name=name,
                nickname=student_data.get('nickname', '').strip() or existing_student.nickname or None,
                photo_path=student_data.get('photo_path', '').strip() or existing_student.photo_path or None,
                cut_count=existing_student.cut_count,
                called_count=existing_student.called_count,
            )
        else:
            # 新增学生
            return Student(
                student_id=student_id,
                name=name,
                nickname=student_data.get('nickname', '').strip() or None,
                photo_path=student_data.get('photo_path', '').strip() or None,
                cut_count=0,
                called_count=0,
            )
    
    def _map_column_name(self, column_name: str) -> Optional[str]:
        """
        将列名映射到标准字段名
        
        Args:
            column_name: 原始列名
            
        Returns:
            标准字段名，如果无法映射则返回None
        """
        column_normalized = column_name.strip()
        column_lower = column_normalized.lower()
        
        for field_name, aliases in STUDENT_COLUMN_MAPPING.items():
            if column_normalized in aliases or column_lower in [a.lower() for a in aliases]:
                return field_name
        return None
    
    def _parse_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """解析CSV文件"""
        from utils.config import Config
        config = Config()
        
        students = []
        with open(file_path, 'r', encoding=config.STUDENT_IMPORT_ENCODING) as f:
            reader = csv.DictReader(f)
            for row in reader:
                student_data = {}
                for key, value in row.items():
                    mapped_field = self._map_column_name(key)
                    if mapped_field:
                        student_data[mapped_field] = value
                students.append(student_data)
        return students
    
    def _parse_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """解析Excel文件"""
        if not HAS_OPENPYXL:
            raise ImportError("需要安装openpyxl库: pip install openpyxl")
        
        students = []
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # 读取表头
        headers = []
        header_row = ws[1]
        for cell in header_row:
            headers.append(cell.value or '')
        
        # 读取数据行
        for row in ws.iter_rows(min_row=2, values_only=False):
            student_data = {}
            for idx, cell in enumerate(row):
                if idx >= len(headers):
                    break
                header_name = str(headers[idx]) if headers[idx] else ''
                value = cell.value
                
                mapped_field = self._map_column_name(header_name)
                if mapped_field:
                    student_data[mapped_field] = str(value) if value else ''
            
            if student_data:
                students.append(student_data)
        
        return students
    
    def _parse_json(self, file_path: str) -> List[Dict[str, Any]]:
        """解析JSON文件"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # 支持数组格式
        if isinstance(data, list):
            return data
        # 支持对象格式，包含students字段
        elif isinstance(data, dict) and 'students' in data:
            return data['students']
        else:
            raise ValueError("JSON格式错误: 期望数组或包含'students'字段的对象")