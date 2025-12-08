#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
XLSX导出器
"""

import os
from typing import Any, Dict, Optional

from services.code_statistics.exporters.base_exporter import Exporter
from models.code_statistics import (
    Summary,
    PythonFunctionStats,
    CFunctionStats,
)


class XLSXExporter(Exporter):
    """XLSX导出器"""
    
    def export(
        self,
        target_dir: str,
        summary: Summary,
        by_language: Dict[str, Summary],
        elapsed_time: float,
        include_comment: bool,
        include_blank: bool,
        function_stats: Optional[PythonFunctionStats],
        c_function_stats: Optional[CFunctionStats],
        detail_table: Optional[Dict[str, Any]],
        base_filename: str,
        update_text_callback: Optional[callable] = None,
    ) -> Optional[str]:
        """导出为XLSX格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            if update_text_callback:
                update_text_callback("保存 XLSX 文件需要 openpyxl 库，请运行: pip install openpyxl\n")
            return None
        
        try:
            xlsx_filename = f"{base_filename}.xlsx"
            xlsx_path = os.path.join(target_dir, xlsx_filename)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "代码统计"
            
            # 设置标题行
            ws["A1"] = "统计项"
            ws["B1"] = "数值"
            ws["A1"].font = Font(bold=True)
            ws["B1"].font = Font(bold=True)
            
            # 写入汇总数据
            row = 2
            ws[f"A{row}"] = "统计目录"
            ws[f"B{row}"] = target_dir
            row += 1
            ws[f"A{row}"] = "总文件数"
            ws[f"B{row}"] = summary.files
            row += 1
            ws[f"A{row}"] = "总行数"
            ws[f"B{row}"] = summary.total
            row += 1
            ws[f"A{row}"] = "代码行数"
            ws[f"B{row}"] = summary.code
            row += 1
            if include_comment:
                ws[f"A{row}"] = "注释行数"
                ws[f"B{row}"] = summary.comment
                row += 1
            if include_blank:
                ws[f"A{row}"] = "空行数"
                ws[f"B{row}"] = summary.blank
                row += 1
            ws[f"A{row}"] = "耗时(秒)"
            ws[f"B{row}"] = f"{elapsed_time:.3f}"
            row += 2
            
            # 按语言统计表
            headers = ["语言", "文件数", "代码行数"]
            if include_comment:
                headers.append("注释行数")
            if include_blank:
                headers.append("空行数")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
            row += 1
            
            for lang, stat in sorted(by_language.items(), key=lambda x: -x[1].code):
                ws.cell(row=row, column=1, value=lang)
                ws.cell(row=row, column=2, value=stat.files)
                ws.cell(row=row, column=3, value=stat.code)
                col_idx = 4
                if include_comment:
                    ws.cell(row=row, column=col_idx, value=stat.comment)
                    col_idx += 1
                if include_blank:
                    ws.cell(row=row, column=col_idx, value=stat.blank)
                row += 1
            
            # Python函数统计
            if function_stats and function_stats.total_functions > 0:
                row += 1
                ws.cell(row=row, column=1, value="Python函数统计").font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=1, value="总函数数")
                ws.cell(row=row, column=2, value=function_stats.total_functions)
                row += 1
                ws.cell(row=row, column=1, value="平均长度")
                ws.cell(row=row, column=2, value=f"{function_stats.mean_length:.2f}")
                row += 1
                ws.cell(row=row, column=1, value="中位数长度")
                ws.cell(row=row, column=2, value=f"{function_stats.median_length:.2f}")
                row += 1
                ws.cell(row=row, column=1, value="最小长度")
                ws.cell(row=row, column=2, value=function_stats.min_length)
                row += 1
                ws.cell(row=row, column=1, value="最大长度")
                ws.cell(row=row, column=2, value=function_stats.max_length)
            
            # C/C++函数统计
            if c_function_stats and c_function_stats.total_functions > 0:
                row += 1
                ws.cell(row=row, column=1, value="C/C++函数统计").font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=1, value="总函数数")
                ws.cell(row=row, column=2, value=c_function_stats.total_functions)
                row += 1
                ws.cell(row=row, column=1, value="平均长度")
                ws.cell(row=row, column=2, value=f"{c_function_stats.mean_length:.2f}")
                row += 1
                ws.cell(row=row, column=1, value="中位数长度")
                ws.cell(row=row, column=2, value=f"{c_function_stats.median_length:.2f}")
                row += 1
                ws.cell(row=row, column=1, value="最小长度")
                ws.cell(row=row, column=2, value=c_function_stats.min_length)
                row += 1
                ws.cell(row=row, column=1, value="最大长度")
                ws.cell(row=row, column=2, value=c_function_stats.max_length)
            
            # 明细表
            if detail_table and detail_table.get("rows"):
                row += 1
                ws_detail = wb.create_sheet("语言明细表")
                for col_idx, header in enumerate(detail_table["columns"], 1):
                    cell = ws_detail.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                for r_idx, values in enumerate(detail_table["rows"], start=2):
                    for c_idx, value in enumerate(values, start=1):
                        ws_detail.cell(row=r_idx, column=c_idx, value=value)
            
            wb.save(xlsx_path)
            return xlsx_filename
        except Exception as exc:
            if update_text_callback:
                update_text_callback(f"保存 XLSX 文件失败: {str(exc)}\n")
            return None
    
    def get_file_extension(self) -> str:
        return "xlsx"

