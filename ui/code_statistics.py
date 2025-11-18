#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
代码统计相关的 Tk 配置窗口与后台处理逻辑。
"""

from __future__ import annotations

import csv
import json
import os
import threading
import time
from datetime import datetime
from typing import Any, Dict, Optional, Set

import tkinter as tk
from tkinter import filedialog, messagebox

from ui.chart_renderer import ChartRenderer


class CodeStatisticsUI:
    """负责代码统计配置窗口、统计执行、结果展示。"""

    DETAIL_LANGUAGES = ("C", "C++", "Java", "Python", "C#")

    def __init__(
        self,
        tk_root: tk.Misc,
        code_counter,
        ui_queue,
        update_text_callback,
        trigger_behavior_callback,
        default_target_dir: Optional[str] = None,
        chart_renderer: Optional[ChartRenderer] = None,
    ):
        self._root = tk_root
        self.code_counter = code_counter
        self._ui_queue = ui_queue
        self._update_text = update_text_callback
        self._trigger_behavior = trigger_behavior_callback
        self._default_target_dir = default_target_dir or os.getcwd()
        self._chart_renderer = chart_renderer or ChartRenderer()

        # 运行时变量
        self._config_window = None
        self._code_counting_config: Optional[Dict[str, Any]] = None
        self._language_vars: Dict[str, tk.BooleanVar] = {}
        self._detail_language_vars: Dict[str, tk.BooleanVar] = {}
        self._target_dir_var: Optional[tk.StringVar] = None

    # ------------------------------------------------------------------ UI --
    def show_config_dialog(self):
        """显示代码统计配置对话框"""
        if not self._root:
            print("错误: Tkinter root窗口未初始化")
            return

        try:
            config_window = tk.Toplevel(self._root)
            config_window.title("代码统计配置")
            config_window.geometry("700x650")
            config_window.minsize(400, 400)
            config_window.focus_set()

            if not self._code_counting_config:
                self._code_counting_config = {
                    "target_dir": self._default_target_dir,
                    "selected_languages": set(),
                    "include_blank": tk.BooleanVar(self._root, value=True),
                    "include_comment": tk.BooleanVar(self._root, value=True),
                    "include_function_stats": tk.BooleanVar(self._root, value=True),
                    "include_c_function_stats": tk.BooleanVar(self._root, value=False),
                    "save_not": tk.BooleanVar(self._root, value=True),
                    "save_csv": tk.BooleanVar(self._root, value=False),
                    "save_json": tk.BooleanVar(self._root, value=False),
                    "save_xlsx": tk.BooleanVar(self._root, value=False),
                }

            self._detail_language_vars = {
                lang: tk.BooleanVar(self._root, value=False) for lang in self.DETAIL_LANGUAGES
            }

            main_frame = tk.Frame(config_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

            # --- 目录选择 ---
            dir_frame = tk.LabelFrame(main_frame, text="选择统计目录", font=("Arial", 11, "bold"), padx=10, pady=10)
            dir_frame.pack(fill=tk.X, padx=5, pady=5)

            dir_input_frame = tk.Frame(dir_frame)
            dir_input_frame.pack(fill=tk.X, padx=5, pady=5)

            self._target_dir_var = tk.StringVar(self._root, value=self._code_counting_config["target_dir"])
            dir_entry = tk.Entry(dir_input_frame, textvariable=self._target_dir_var, font=("Arial", 10), width=50)
            dir_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))

            def browse_directory():
                directory = filedialog.askdirectory(title="选择要统计的目录", initialdir=self._target_dir_var.get())
                if directory:
                    self._target_dir_var.set(directory)
                    self._code_counting_config["target_dir"] = directory

            tk.Button(dir_input_frame, text="浏览...", command=browse_directory, font=("Arial", 10), width=10).pack(
                side=tk.RIGHT
            )

            # --- 语言选择 ---
            lang_frame = tk.LabelFrame(
                main_frame, text="选择要统计的语言（不选择则统计所有语言）", font=("Arial", 11, "bold"), padx=10, pady=10
            )
            lang_frame.pack(fill=tk.BOTH, expand=False, padx=5, pady=5)

            all_languages = sorted(set(self.code_counter.EXT_TO_LANGUAGE.values()))
            self._language_vars = {}

            lang_canvas = tk.Canvas(lang_frame, height=180)
            lang_scrollbar = tk.Scrollbar(lang_frame, orient="vertical", command=lang_canvas.yview)
            lang_scrollable_frame = tk.Frame(lang_canvas)

            _last_update_time = [0]

            def update_scrollregion(event=None):
                try:
                    current_time = time.time()
                    if current_time - _last_update_time[0] < 0.1:
                        return
                    _last_update_time[0] = current_time
                    lang_canvas.configure(scrollregion=lang_canvas.bbox("all"))
                except Exception:
                    pass

            lang_scrollable_frame.bind("<Configure>", update_scrollregion)

            lang_canvas_window_id = lang_canvas.create_window((0, 0), window=lang_scrollable_frame, anchor="nw")
            lang_canvas.configure(yscrollcommand=lang_scrollbar.set)

            def on_canvas_configure(event):
                try:
                    canvas_width = event.width
                    lang_canvas.itemconfig(lang_canvas_window_id, width=canvas_width)
                    update_scrollregion()
                except Exception:
                    pass

            lang_canvas.bind("<Configure>", on_canvas_configure)

            cols = 3
            for idx, lang in enumerate(all_languages):
                var = tk.BooleanVar(self._root, value=False)
                self._language_vars[lang] = var
                row = idx // cols
                col = idx % cols
                tk.Checkbutton(lang_scrollable_frame, text=lang, variable=var, font=("Arial", 9)).grid(
                    row=row, column=col, sticky=tk.W, padx=5, pady=2
                )

            lang_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            lang_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            lang_btn_frame = tk.Frame(lang_frame)
            lang_btn_frame.pack(padx=5, pady=5)

            def select_all_languages():
                for var in self._language_vars.values():
                    var.set(True)

            def deselect_all_languages():
                for var in self._language_vars.values():
                    var.set(False)

            def toggle_select_all(event=None):
                if all(var.get() for var in self._language_vars.values()):
                    deselect_all_languages()
                else:
                    select_all_languages()

            tk.Button(lang_btn_frame, text="全选", command=toggle_select_all, font=("Arial", 10), width=10).pack(
                side=tk.LEFT
            )
            tk.Button(lang_btn_frame, text="全不选", command=deselect_all_languages, font=("Arial", 10), width=10).pack(
                side=tk.LEFT
            )

            # --- 统计选项 ---
            options_frame = tk.LabelFrame(main_frame, text="统计选项", font=("Arial", 11, "bold"), padx=10, pady=10)
            options_frame.pack(fill=tk.X, padx=5, pady=5)

            tk.Checkbutton(
                options_frame,
                text="统计空行数",
                variable=self._code_counting_config["include_blank"],
                font=("Arial", 10),
            ).pack(anchor=tk.W, padx=10, pady=5)
            tk.Checkbutton(
                options_frame,
                text="统计注释行数",
                variable=self._code_counting_config["include_comment"],
                font=("Arial", 10),
            ).pack(anchor=tk.W, padx=10, pady=5)
            tk.Checkbutton(
                options_frame,
                text="统计Python函数信息（仅对Python文件）",
                variable=self._code_counting_config["include_function_stats"],
                font=("Arial", 10),
            ).pack(anchor=tk.W, padx=10, pady=5)
            tk.Checkbutton(
                options_frame,
                text="统计C/C++函数信息（仅对C/C++文件）",
                variable=self._code_counting_config["include_c_function_stats"],
                font=("Arial", 10),
            ).pack(anchor=tk.W, padx=10, pady=5)

            # --- 保存选项 ---
            save_frame = tk.LabelFrame(main_frame, text="保存选项", font=("Arial", 11, "bold"), padx=10, pady=10)
            save_frame.pack(fill=tk.X, padx=5, pady=5)

            save_not_cb = tk.Checkbutton(
                save_frame,
                text="不保存",
                variable=self._code_counting_config["save_not"],
                font=("Arial", 10),
            )
            save_not_cb.pack(anchor=tk.W, padx=10, pady=5)
            save_csv_cb = tk.Checkbutton(
                save_frame, text="保存为 CSV", variable=self._code_counting_config["save_csv"], font=("Arial", 10)
            )
            save_csv_cb.pack(anchor=tk.W, padx=10, pady=5)
            save_json_cb = tk.Checkbutton(
                save_frame, text="保存为 JSON", variable=self._code_counting_config["save_json"], font=("Arial", 10)
            )
            save_json_cb.pack(anchor=tk.W, padx=10, pady=5)
            save_xlsx_cb = tk.Checkbutton(
                save_frame, text="保存为 XLSX", variable=self._code_counting_config["save_xlsx"], font=("Arial", 10)
            )
            save_xlsx_cb.pack(anchor=tk.W, padx=10, pady=5)

            def update_save_options():
                if self._code_counting_config["save_not"].get():
                    for widget in (save_csv_cb, save_json_cb, save_xlsx_cb):
                        widget.config(state=tk.DISABLED)
                    self._code_counting_config["save_csv"].set(False)
                    self._code_counting_config["save_json"].set(False)
                    self._code_counting_config["save_xlsx"].set(False)
                else:
                    for widget in (save_csv_cb, save_json_cb, save_xlsx_cb):
                        widget.config(state=tk.NORMAL)

            self._code_counting_config["save_not"].trace("w", lambda *args: update_save_options())
            update_save_options()

            # --- 语言明细表 ---
            detail_frame = tk.LabelFrame(
                main_frame,
                text="语言明细表（勾选后在图表展示并可随导出）",
                font=("Arial", 11, "bold"),
                padx=10,
                pady=10,
            )
            detail_frame.pack(fill=tk.X, padx=5, pady=5)
            tk.Label(detail_frame, text="请选择需要生成详细表格的语言：", font=("Arial", 9)).pack(anchor=tk.W, padx=5, pady=(0, 5))
            detail_container = tk.Frame(detail_frame)
            detail_container.pack(fill=tk.X)
            for idx, lang in enumerate(self.DETAIL_LANGUAGES):
                tk.Checkbutton(
                    detail_container, text=lang, variable=self._detail_language_vars[lang], font=("Arial", 10)
                ).grid(row=0, column=idx, padx=5, pady=2, sticky=tk.W)

            # --- 按钮 ---
            button_frame = tk.Frame(main_frame)
            button_frame.pack(fill=tk.X, padx=5, pady=5, side=tk.BOTTOM)

            def cleanup_config_dialog():
                try:
                    if self._target_dir_var:
                        self._target_dir_var = None
                    if self._language_vars:
                        self._language_vars.clear()
                    if self._code_counting_config:
                        for key in list(self._code_counting_config.keys()):
                            if isinstance(self._code_counting_config[key], (tk.BooleanVar, tk.StringVar)):
                                self._code_counting_config[key] = None
                            else:
                                del self._code_counting_config[key]
                        self._code_counting_config = None
                    self._config_window = None
                except Exception:
                    pass
                config_window.destroy()

            def start_counting():
                target_dir = self._target_dir_var.get().strip() if self._target_dir_var else ""
                if not target_dir or not os.path.exists(target_dir):
                    messagebox.showerror("错误", "请选择有效的目录路径！")
                    return

                selected_languages = {lang for lang, var in self._language_vars.items() if var.get()}
                include_blank = self._code_counting_config["include_blank"].get()
                include_comment = self._code_counting_config["include_comment"].get()
                include_function_stats = self._code_counting_config["include_function_stats"].get()
                include_c_function_stats = self._code_counting_config["include_c_function_stats"].get()
                save_not = self._code_counting_config["save_not"].get()
                save_csv = self._code_counting_config["save_csv"].get()
                save_json = self._code_counting_config["save_json"].get()
                save_xlsx = self._code_counting_config["save_xlsx"].get()
                detail_languages = {lang for lang, var in self._detail_language_vars.items() if var.get()}

                self._update_text(f"唐老鸭: 好的！开始统计代码量！\n目录: {target_dir}\n")
                if selected_languages:
                    lang_list = ", ".join(sorted(selected_languages))
                    self._update_text(f"统计语言: {lang_list}\n\n")

                threading.Thread(
                    target=self.start_code_counting,
                    args=(
                        target_dir,
                        selected_languages,
                        include_blank,
                        include_comment,
                        include_function_stats,
                        include_c_function_stats,
                        save_not,
                        save_csv,
                        save_json,
                        save_xlsx,
                        detail_languages,
                    ),
                    daemon=True,
                ).start()

            tk.Button(
                button_frame,
                text="开始统计",
                command=start_counting,
                font=("Arial", 10),
                width=10,
                bg="#4CAF50",
                fg="white",
            ).pack(side=tk.LEFT, padx=5)

            tk.Button(button_frame, text="取消", command=cleanup_config_dialog, font=("Arial", 10), width=10).pack(
                side=tk.LEFT, padx=5
            )
            tk.Frame(button_frame).pack(side=tk.LEFT, fill=tk.X, expand=True)

            config_window.protocol("WM_DELETE_WINDOW", cleanup_config_dialog)
            config_window.update_idletasks()
            config_window.focus_set()
            self._config_window = config_window
        except Exception as exc:
            print(f"创建代码统计配置对话框时出错: {exc}")
            import traceback

            traceback.print_exc()
            messagebox.showerror("错误", f"创建配置对话框失败: {exc}")

    # --------------------------------------------------------------- BACKEND --
    def start_code_counting(
        self,
        target_dir: Optional[str] = None,
        selected_languages: Optional[Set[str]] = None,
        include_blank: bool = True,
        include_comment: bool = True,
        include_function_stats: bool = True,
        include_c_function_stats: bool = False,
        save_not: bool = True,
        save_csv: bool = False,
        save_json: bool = False,
        save_xlsx: bool = False,
        detail_languages: Optional[Set[str]] = None,
    ):
        """启动代码统计（后台线程）"""
        detail_languages = detail_languages or set()
        self._ui_queue.put(("change_duckling_theme", "focused"))
        self._trigger_behavior("code_count")

        try:
            self._update_text("唐老鸭: 正在统计代码量，请稍候...\n")

            if target_dir is None:
                target_dir = self._default_target_dir

            include_patterns = []
            if selected_languages:
                ext_to_lang = self.code_counter.EXT_TO_LANGUAGE
                lang_to_exts: Dict[str, list] = {}
                for ext, lang in ext_to_lang.items():
                    lang_to_exts.setdefault(lang, []).append(ext)
                for lang in selected_languages:
                    for ext in lang_to_exts.get(lang, []):
                        include_patterns.append(f"**/*{ext}")

            result = self.code_counter.count_code_by_language(
                target_dir, include=include_patterns if include_patterns else None
            )
            summary = result["summary"]
            by_language = result["by_language"]
            elapsed_time = result["elapsed_time"]

            if selected_languages:
                by_language = {lang: stat for lang, stat in by_language.items() if lang in selected_languages}

            report_lines = [
                "代码统计报告:",
                "=" * 50,
                f"统计目录: {target_dir}",
                f"总文件数: {summary.files}",
                f"总行数: {summary.total}",
                f"代码行数: {summary.code}",
            ]
            if include_comment:
                report_lines.append(f"注释行数: {summary.comment}")
            if include_blank:
                report_lines.append(f"空行数: {summary.blank}")
            report_lines.append(f"耗时: {elapsed_time:.3f} 秒\n")

            if by_language:
                header = f"{'语言':<20} {'文件数':<10} {'代码行数':<15}"
                if include_comment:
                    header += f" {'注释行数':<15}"
                if include_blank:
                    header += f" {'空行数':<15}"
                report_lines.append("按语言统计:")
                report_lines.append(header)
                report_lines.append("-" * 80)
                for lang, stat in sorted(by_language.items(), key=lambda x: -x[1].code):
                    row = f"{lang:<20} {stat.files:<10} {stat.code:<15}"
                    if include_comment:
                        row += f" {stat.comment:<15}"
                    if include_blank:
                        row += f" {stat.blank:<15}"
                    report_lines.append(row)
            else:
                report_lines.append("未找到匹配的代码文件。")

            function_stats = None
            if include_function_stats:
                function_stats = self.code_counter.count_python_functions(target_dir)
                if function_stats.total_functions > 0:
                    report_lines.extend(
                        [
                            "",
                            "Python函数统计:",
                            f"总函数数: {function_stats.total_functions}",
                            f"平均长度: {function_stats.mean_length:.2f} 行",
                            f"中位数长度: {function_stats.median_length:.2f} 行",
                            f"最小长度: {function_stats.min_length} 行",
                            f"最大长度: {function_stats.max_length} 行",
                        ]
                    )

            c_function_stats = None
            has_c_like_language = any(
                lang.lower() in {"c", "c++", "c/c++ header", "c++ header"} for lang in by_language.keys()
            )
            if include_c_function_stats or has_c_like_language:
                c_function_stats = self.code_counter.count_c_functions(target_dir)
                if include_c_function_stats and c_function_stats.total_functions > 0:
                    report_lines.extend(
                        [
                            "",
                            "C/C++函数统计:",
                            f"总函数数: {c_function_stats.total_functions}",
                            f"平均长度: {c_function_stats.mean_length:.2f} 行",
                            f"中位数长度: {c_function_stats.median_length:.2f} 行",
                            f"最小长度: {c_function_stats.min_length} 行",
                            f"最大长度: {c_function_stats.max_length} 行",
                        ]
                    )

            detail_table = self._build_detail_table_data(
                by_language, detail_languages, include_blank, include_comment, function_stats, c_function_stats
            )

            self._update_text(f"唐老鸭: 代码统计完成！\n" + "\n".join(report_lines) + "\n")

            if not save_not:
                self._save_results(
                    target_dir,
                    summary,
                    by_language,
                    elapsed_time,
                    include_comment,
                    include_blank,
                    function_stats,
                    c_function_stats,
                    detail_table,
                    save_csv,
                    save_json,
                    save_xlsx,
                )

            self._enqueue_show_charts(result, function_stats, c_function_stats, detail_table)
            self._ui_queue.put(("change_duckling_theme", "original"))
        except Exception as exc:
            print(f"代码统计错误: {exc}")
            import traceback

            traceback.print_exc()
            self._update_text(f"唐老鸭: 抱歉，代码统计出现了问题: {str(exc)}\n\n")
            self._ui_queue.put(("change_duckling_theme", "original"))

    # ------------------------------------------------------------- HELPERS --
    def _build_detail_table_data(
        self,
        by_language,
        detail_languages,
        include_blank,
        include_comment,
        function_stats,
        c_function_stats,
    ):
        if not detail_languages:
            return None

        columns = ["语言", "源文件数", "代码行数", "空行数", "注释行数", "函数个数", "最大值", "最小值", "均值", "中位数"]
        rows = []

        def match_lang(key):
            for existing in by_language.keys():
                if existing.lower() == key.lower():
                    return existing
            return None

        def extract_metrics(lang_name):
            metrics = {"count": 0, "min": "-", "max": "-", "mean": "-", "median": "-"}
            lower = lang_name.lower()
            if "python" in lower and function_stats:
                metrics["count"] = getattr(function_stats, "total_functions", 0) or 0
                metrics["min"] = getattr(function_stats, "min_length", 0) or 0
                metrics["max"] = getattr(function_stats, "max_length", 0) or 0
                metrics["mean"] = round(getattr(function_stats, "mean_length", 0.0) or 0.0, 2)
                metrics["median"] = round(getattr(function_stats, "median_length", 0.0) or 0.0, 2)
            elif lower in {"c", "c++", "c/c++"} and c_function_stats:
                metrics["count"] = getattr(c_function_stats, "total_functions", 0) or 0
                metrics["min"] = getattr(c_function_stats, "min_length", 0) or 0
                metrics["max"] = getattr(c_function_stats, "max_length", 0) or 0
                metrics["mean"] = round(getattr(c_function_stats, "mean_length", 0.0) or 0.0, 2)
                metrics["median"] = round(getattr(c_function_stats, "median_length", 0.0) or 0.0, 2)
            else:
                metrics["count"] = 0
            return metrics

        for lang in sorted(detail_languages):
            lang_key = match_lang(lang) or lang
            stat = by_language.get(lang_key)
            file_count = getattr(stat, "files", 0) if stat else 0
            code_lines = getattr(stat, "code", 0) if stat else 0
            comment_lines = getattr(stat, "comment", 0) if stat else 0
            blank_lines = getattr(stat, "blank", 0) if stat else 0
            comment_display = comment_lines if include_comment else "-"
            blank_display = blank_lines if include_blank else "-"

            metrics = extract_metrics(lang_key)
            rows.append(
                [
                    lang,
                    file_count,
                    code_lines,
                    blank_display,
                    comment_display,
                    metrics["count"],
                    metrics["max"],
                    metrics["min"],
                    metrics["mean"],
                    metrics["median"],
                ]
            )

        if not rows:
            return None

        return {"title": "语言明细表", "columns": columns, "rows": rows}

    def _save_results(
        self,
        target_dir,
        summary,
        by_language,
        elapsed_time,
        include_comment,
        include_blank,
        function_stats,
        c_function_stats,
        detail_table,
        save_csv,
        save_json,
        save_xlsx,
    ):
        saved_files = []
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_filename = f"code_statistics_{timestamp}"

        save_data = {
            "summary": {
                "target_dir": target_dir,
                "files": summary.files,
                "total": summary.total,
                "code": summary.code,
                "comment": summary.comment if include_comment else None,
                "blank": summary.blank if include_blank else None,
                "elapsed_time": elapsed_time,
            },
            "by_language": {},
        }

        for lang, stat in by_language.items():
            lang_data = {"files": stat.files, "code": stat.code}
            if include_comment:
                lang_data["comment"] = stat.comment
            if include_blank:
                lang_data["blank"] = stat.blank
            save_data["by_language"][lang] = lang_data

        if function_stats and function_stats.total_functions > 0:
            save_data["python_functions"] = {
                "total_functions": function_stats.total_functions,
                "mean_length": function_stats.mean_length,
                "median_length": function_stats.median_length,
                "min_length": function_stats.min_length,
                "max_length": function_stats.max_length,
            }

        if c_function_stats and c_function_stats.total_functions > 0:
            save_data["c_functions"] = {
                "total_functions": c_function_stats.total_functions,
                "mean_length": c_function_stats.mean_length,
                "median_length": c_function_stats.median_length,
                "min_length": c_function_stats.min_length,
                "max_length": c_function_stats.max_length,
            }

        if detail_table and detail_table.get("rows"):
            save_data["detail_table"] = detail_table

        if save_csv:
            try:
                csv_filename = f"{base_filename}.csv"
                csv_path = os.path.join(target_dir, csv_filename)
                with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(["统计项", "数值"])
                    writer.writerow(["统计目录", target_dir])
                    writer.writerow(["总文件数", summary.files])
                    writer.writerow(["总行数", summary.total])
                    writer.writerow(["代码行数", summary.code])
                    if include_comment:
                        writer.writerow(["注释行数", summary.comment])
                    if include_blank:
                        writer.writerow(["空行数", summary.blank])
                    writer.writerow(["耗时(秒)", f"{elapsed_time:.3f}"])
                    writer.writerow([])
                    header = ["语言", "文件数", "代码行数"]
                    if include_comment:
                        header.append("注释行数")
                    if include_blank:
                        header.append("空行数")
                    writer.writerow(header)
                    for lang, stat in sorted(by_language.items(), key=lambda x: -x[1].code):
                        row = [lang, stat.files, stat.code]
                        if include_comment:
                            row.append(stat.comment)
                        if include_blank:
                            row.append(stat.blank)
                        writer.writerow(row)

                    if function_stats and function_stats.total_functions > 0:
                        writer.writerow([])
                        writer.writerow(["Python函数统计"])
                        writer.writerow(["总函数数", function_stats.total_functions])
                        writer.writerow(["平均长度", f"{function_stats.mean_length:.2f}"])
                        writer.writerow(["中位数长度", f"{function_stats.median_length:.2f}"])
                        writer.writerow(["最小长度", function_stats.min_length])
                        writer.writerow(["最大长度", function_stats.max_length])

                    if c_function_stats and c_function_stats.total_functions > 0:
                        writer.writerow([])
                        writer.writerow(["C/C++函数统计"])
                        writer.writerow(["总函数数", c_function_stats.total_functions])
                        writer.writerow(["平均长度", f"{c_function_stats.mean_length:.2f}"])
                        writer.writerow(["中位数长度", f"{c_function_stats.median_length:.2f}"])
                        writer.writerow(["最小长度", c_function_stats.min_length])
                        writer.writerow(["最大长度", c_function_stats.max_length])

                    if detail_table and detail_table.get("rows"):
                        writer.writerow([])
                        writer.writerow(["语言明细表"])
                        writer.writerow(detail_table["columns"])
                        for row in detail_table["rows"]:
                            writer.writerow(row)

                saved_files.append(csv_filename)
            except Exception as exc:
                self._update_text(f"保存 CSV 文件失败: {str(exc)}\n")

        if save_json:
            try:
                json_filename = f"{base_filename}.json"
                json_path = os.path.join(target_dir, json_filename)
                with open(json_path, "w", encoding="utf-8") as f:
                    json.dump(save_data, f, ensure_ascii=False, indent=2)
                saved_files.append(json_filename)
            except Exception as exc:
                self._update_text(f"保存 JSON 文件失败: {str(exc)}\n")

        if save_xlsx:
            try:
                import openpyxl
                from openpyxl.styles import Font
            except ImportError:
                self._update_text("保存 XLSX 文件需要 openpyxl 库，请运行: pip install openpyxl\n")
            else:
                try:
                    xlsx_filename = f"{base_filename}.xlsx"
                    xlsx_path = os.path.join(target_dir, xlsx_filename)
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "代码统计"

                    ws["A1"] = "统计项"
                    ws["B1"] = "数值"
                    ws["A1"].font = Font(bold=True)
                    ws["B1"].font = Font(bold=True)

                    row = 2
                    ws[f"A{row}"] = "统计目录"
                    ws[f"B{row}"] = target_dir
                    row += 1
                    ws[f"A{row}"] = "总文件数"
                    ws[f"B{row}"] = summary.files
                    row += 1
                    ws[f"A{row}"] = "总行数"
                    ws[f"B{row}"] = summary.total
                    row += 1
                    ws[f"A{row}"] = "代码行数"
                    ws[f"B{row}"] = summary.code
                    row += 1
                    if include_comment:
                        ws[f"A{row}"] = "注释行数"
                        ws[f"B{row}"] = summary.comment
                        row += 1
                    if include_blank:
                        ws[f"A{row}"] = "空行数"
                        ws[f"B{row}"] = summary.blank
                        row += 1
                    ws[f"A{row}"] = "耗时(秒)"
                    ws[f"B{row}"] = f"{elapsed_time:.3f}"
                    row += 2

                    headers = ["语言", "文件数", "代码行数"]
                    if include_comment:
                        headers.append("注释行数")
                    if include_blank:
                        headers.append("空行数")

                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=row, column=col, value=header)
                        cell.font = Font(bold=True)
                    row += 1

                    for lang, stat in sorted(by_language.items(), key=lambda x: -x[1].code):
                        ws.cell(row=row, column=1, value=lang)
                        ws.cell(row=row, column=2, value=stat.files)
                        ws.cell(row=row, column=3, value=stat.code)
                        col_idx = 4
                        if include_comment:
                            ws.cell(row=row, column=col_idx, value=stat.comment)
                            col_idx += 1
                        if include_blank:
                            ws.cell(row=row, column=col_idx, value=stat.blank)
                        row += 1

                    if function_stats and function_stats.total_functions > 0:
                        row += 1
                        ws.cell(row=row, column=1, value="Python函数统计").font = Font(bold=True)
                        row += 1
                        ws.cell(row=row, column=1, value="总函数数")
                        ws.cell(row=row, column=2, value=function_stats.total_functions)
                        row += 1
                        ws.cell(row=row, column=1, value="平均长度")
                        ws.cell(row=row, column=2, value=f"{function_stats.mean_length:.2f}")
                        row += 1
                        ws.cell(row=row, column=1, value="中位数长度")
                        ws.cell(row=row, column=2, value=f"{function_stats.median_length:.2f}")
                        row += 1
                        ws.cell(row=row, column=1, value="最小长度")
                        ws.cell(row=row, column=2, value=function_stats.min_length)
                        row += 1
                        ws.cell(row=row, column=1, value="最大长度")
                        ws.cell(row=row, column=2, value=function_stats.max_length)

                    if c_function_stats and c_function_stats.total_functions > 0:
                        row += 1
                        ws.cell(row=row, column=1, value="C/C++函数统计").font = Font(bold=True)
                        row += 1
                        ws.cell(row=row, column=1, value="总函数数")
                        ws.cell(row=row, column=2, value=c_function_stats.total_functions)
                        row += 1
                        ws.cell(row=row, column=1, value="平均长度")
                        ws.cell(row=row, column=2, value=f"{c_function_stats.mean_length:.2f}")
                        row += 1
                        ws.cell(row=row, column=1, value="中位数长度")
                        ws.cell(row=row, column=2, value=f"{c_function_stats.median_length:.2f}")
                        row += 1
                        ws.cell(row=row, column=1, value="最小长度")
                        ws.cell(row=row, column=2, value=c_function_stats.min_length)
                        row += 1
                        ws.cell(row=row, column=1, value="最大长度")
                        ws.cell(row=row, column=2, value=c_function_stats.max_length)

                    if detail_table and detail_table.get("rows"):
                        row += 1
                        ws_detail = wb.create_sheet("语言明细表")
                        for col_idx, header in enumerate(detail_table["columns"], 1):
                            cell = ws_detail.cell(row=1, column=col_idx, value=header)
                            cell.font = Font(bold=True)
                        for r_idx, values in enumerate(detail_table["rows"], start=2):
                            for c_idx, value in enumerate(values, start=1):
                                ws_detail.cell(row=r_idx, column=c_idx, value=value)

                    wb.save(xlsx_path)
                    saved_files.append(xlsx_filename)
                except Exception as exc:
                    self._update_text(f"保存 XLSX 文件失败: {str(exc)}\n")

        if saved_files:
            files_list = ", ".join(saved_files)
            self._update_text(f"统计结果已保存: {files_list}\n保存位置: {target_dir}\n\n")

    # ---------------------------------------------------------- UI QUEUE OPS --
    def _enqueue_show_charts(self, code_result, function_stats=None, c_function_stats=None, detail_table=None):
        try:
            self._ui_queue.put(("show_charts", code_result, function_stats, c_function_stats, detail_table), block=False)
        except Exception as exc:
            print(f"提交图表显示到队列失败: {exc}")

    def show_charts(self, code_result, function_stats=None, c_function_stats=None, detail_table=None):
        """供主线程调用，真正绘制图表。"""
        self._chart_renderer.show_code_statistics_charts(code_result, function_stats, c_function_stats, detail_table)


